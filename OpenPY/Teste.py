from openpyxl import Workbook

fieldnames = [
			  'Data' , 
			  'Numero chamado' , 
			  'NIC da Chamada',
			  'Fila do Contact Service',
			  'Nome do Agente',
			  'Habilidade da chamada',
			  'Hora Inicio da chamada',
			  'Hora de Abandono da chamada',
			  'Calculo de horas'
			  ]

fieldados = ['2018/10/2018','1','7583','Skill','Lucas','TOP','22h','21h','23h','10']

dicionario = [{"Data" : "2018/10/2018" ,
			  "Numero chamado" : "8702" ,
			  "NIC da Chamada" : "11977931383",
			  "Fila do Contact Service" : "CSQ_DSADSADA", 
			  "Nome do Agente" : "Lucas",
			  "Habilidade da chamada" : "SKILL FILA FALE",
			  "Hora Inicio da chamada" : "23h",
			  "Hora de Abandono da chamada" : "00",
			  "Calculo de horas" : "7"}];

wb = Workbook()
ws = wb.active
ws = wb.create_sheet("Relatorio Resumo")

i = 1
while(i < 9):
	ws.cell(column=i, row=1,value=fieldnames[i])
	ws.cell(column=i, row=2,value=fieldados[i])
	i = i + 1

wb.remove(wb["Sheet"])
wb.save('/home/lucas/Desktop/TesteLinux.xlsx')